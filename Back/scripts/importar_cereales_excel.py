# -*- coding: utf-8 -*-
"""
Script para importar productos desde Excel
Valida y importa productos de la categoría Cereales
"""

import sys
import os

# Agregar el directorio raíz del backend al path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from app.database import get_db_connection
from app.config import settings

# Configuración
ARCHIVO_EXCEL = r"d:\CBA 4.0\AymaraContable\Nuevo Hoja de cálculo de Microsoft Excel.xlsx"
CATEGORIA_ID = 2  # Cereales
PROVEEDOR_ID = 1  # Proveedor default

# Unidades válidas
UNIDADES_VALIDAS = ['GRAMO', 'KILO', 'BOLSA_XKILO', 'g', 'kg', 'bolsa_xkilo']

def validar_unidad(unidad: str) -> bool:
    """Valida que la unidad sea aceptable"""
    if not unidad:
        return False
    unidad_upper = unidad.upper().strip()
    return unidad_upper in ['GRAMO', 'KILO', 'BOLSA_XKILO', 'G', 'KG'] or unidad_upper.startswith('BOLSA_')

def normalizar_unidad(unidad: str) -> str:
    """Normaliza la unidad a formato estándar"""
    if not unidad:
        return 'GRAMO'
    unidad = unidad.upper().strip()
    # Mapeo de unidades
    if unidad in ['G', 'GRAMO']:
        return 'GRAMO'
    if unidad in ['KG', 'KILO']:
        return 'KILO'
    if unidad.startswith('BOLSA_'):
        return unidad  # Mantener formato BOLSA_X.XKILO
    return 'GRAMO'

def verificar_categoria(conn, categoria_id: int) -> bool:
    """Verifica que la categoría exista"""
    with conn.cursor() as cur:
        cur.execute("SELECT id, nombre FROM categorias WHERE id = %s", (categoria_id,))
        result = cur.fetchone()
        if result:
            print(f"✅ Categoría verificada: ID={result[0]}, Nombre={result[1]}")
            return True
        print(f"❌ Categoría no encontrada: ID={categoria_id}")
        return False

def verificar_proveedor(conn, proveedor_id: int) -> bool:
    """Verifica que el proveedor exista"""
    with conn.cursor() as cur:
        cur.execute("SELECT id, nombre FROM proveedores WHERE id = %s", (proveedor_id,))
        result = cur.fetchone()
        if result:
            print(f"✅ Proveedor verificado: ID={result[0]}, Nombre={result[1]}")
            return True
        print(f"❌ Proveedor no encontrado: ID={proveedor_id}")
        return False

def verificar_sku_duplicado(conn, sku: str) -> bool:
    """Verifica si el SKU ya existe en la base de datos"""
    with conn.cursor() as cur:
        cur.execute("SELECT id, nombre FROM productos WHERE sku = %s", (sku,))
        result = cur.fetchone()
        if result:
            print(f"  ⚠️ SKU duplicado: {sku} (ya existe producto ID={result[0]}, Nombre={result[1]})")
            return True
        return False

def importar_productos():
    """Función principal para importar productos desde Excel"""
    
    print("=" * 60)
    print("IMPORTACIÓN DE PRODUCTOS - CEREALES")
    print("=" * 60)
    
    # Verificar archivo Excel
    if not os.path.exists(ARCHIVO_EXCEL):
        print(f"❌ Error: No se encontró el archivo Excel en: {ARCHIVO_EXCEL}")
        return
    
    print(f"📁 Archivo Excel: {ARCHIVO_EXCEL}")
    
    # Cargar libro Excel
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
        print(f"📊 Hoja activa: {ws.title}")
    except Exception as e:
        print(f"❌ Error al leer Excel: {e}")
        return
    
    # Conectar a base de datos
    conn = None
    try:
        conn = get_db_connection()
        print("✅ Conexión a base de datos exitosa")
        
        # Verificaciones previas
        print("\n--- VERIFICACIONES PREVIAS ---")
        if not verificar_categoria(conn, CATEGORIA_ID):
            print("❌ Abortando: Categoría no existe")
            return
        if not verificar_proveedor(conn, PROVEEDOR_ID):
            print("❌ Abortando: Proveedor no existe")
            return
        
        # Leer datos del Excel
        print("\n--- LEYENDO DATOS DEL EXCEL ---")
        
        # Asumir que la primera fila son encabezados
        # Columnas esperadas: sku, nombre, unidad_medida, precio_costo, precio_venta, stock
        
        errores = []
        productos_validos = []
        
        # Iterar filas (empezando desde fila 2, asumiendo fila 1 = encabezados)
        for fila_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Saltar filas vacías
            if not any(row):
                continue
            
            # Extraer columnas (ajustar índices según estructura real del Excel)
            # Asumimos: A=sku, B=nombre, C=unidad_medida, D=precio_costo, E=precio_venta, F=stock
            sku = str(row[0]).strip() if row[0] else None
            nombre = str(row[1]).strip() if row[1] else None
            unidad_medida = str(row[2]).strip() if row[2] else 'GRAMO'
            precio_costo = float(row[3]) if row[3] else 0
            precio_venta = float(row[4]) if row[4] else 0
            stock = float(row[5]) if row[5] else 0
            
            print(f"\nFila {fila_idx}: {nombre}")
            print(f"  SKU: {sku}, Unidad: {unidad_medida}, Precio: ${precio_venta}")
            
            # Validaciones
            fila_errores = []
            
            # 1. Validar nombre
            if not nombre:
                fila_errores.append("Nombre vacío")
            
            # 2. Validar unidad de medida
            if not validar_unidad(unidad_medida):
                fila_errores.append(f"Unidad inválida: {unidad_medida}")
            
            # 3. Validar SKU duplicado (si tiene SKU)
            if sku and verificar_sku_duplicado(conn, sku):
                fila_errores.append(f"SKU duplicado: {sku}")
            
            # 4. Validar precio
            if precio_venta <= 0:
                fila_errores.append("Precio de venta inválido")
            
            if fila_errores:
                errores.append({
                    'fila': fila_idx,
                    'nombre': nombre,
                    'errores': fila_errores
                })
                print(f"  ❌ Errores: {fila_errores}")
            else:
                productos_validos.append({
                    'fila': fila_idx,
                    'sku': sku,
                    'nombre': nombre,
                    'unidad_medida': normalizar_unidad(unidad_medida),
                    'precio_costo': precio_costo,
                    'precio_venta': precio_venta,
                    'stock': stock
                })
                print(f"  ✅ Válido")
        
        # Reporte de validación
        print("\n" + "=" * 60)
        print("RESUMEN DE VALIDACIÓN")
        print("=" * 60)
        print(f"Total filas leídas: {len(productos_validos) + len(errores)}")
        print(f"Productos válidos: {len(productos_validos)}")
        print(f"Errores encontrados: {len(errores)}")
        
        if errores:
            print("\n--- ERRORES DETALLADOS ---")
            for error in errores:
                print(f"  Fila {error['fila']}: {error['nombre']}")
                for err in error['errores']:
                    print(f"    - {err}")
        
        # Si hay errores, preguntar si continuar
        if errores:
            continuar = input("\n¿Continuar solo con productos válidos? (s/n): ").lower()
            if continuar != 's':
                print("Importación cancelada")
                return
        
        # Importar productos válidos
        if productos_validos:
            print("\n--- IMPORTANDO PRODUCTOS ---")
            
            with conn.cursor() as cur:
                for prod in productos_validos:
                    try:
                        cur.execute("""
                            INSERT INTO productos (
                                sku, nombre, categoria_id, proveedor_id, 
                                unidad_medida, precio_costo, precio_venta, 
                                stock_actual, activo
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TRUE)
                            RETURNING id
                        """, (
                            prod['sku'],
                            prod['nombre'],
                            CATEGORIA_ID,
                            PROVEEDOR_ID,
                            prod['unidad_medida'],
                            prod['precio_costo'],
                            prod['precio_venta'],
                            prod['stock']
                        ))
                        result = cur.fetchone()
                        print(f"  ✅ Importado: {prod['nombre']} (ID={result[0]})")
                    except Exception as e:
                        print(f"  ❌ Error al importar {prod['nombre']}: {e}")
            
            conn.commit()
            print("\n" + "=" * 60)
            print(f"✅ IMPORTACIÓN COMPLETADA")
            print(f"   Productos importados: {len(productos_validos)}")
            print("=" * 60)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()
            print("🔒 Conexión cerrada")

if __name__ == "__main__":
    importar_productos()
