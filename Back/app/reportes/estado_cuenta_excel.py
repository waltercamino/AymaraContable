"""
Generador de Excel para Estado de Cuenta
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from io import BytesIO
from datetime import datetime
from typing import List, Optional


def format_currency(value: float) -> str:
    """Formatea un valor como moneda argentina"""
    return f"${value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def generar_excel(
    entidad_nombre: str,
    entidad_cuit: Optional[str],
    tipo: str,  # 'cliente' o 'proveedor'
    movimientos: List[dict],
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    saldo_final: float = 0
) -> bytes:
    """
    Genera un Excel con el estado de cuenta de un cliente o proveedor.
    
    Args:
        entidad_nombre: Nombre del cliente o proveedor
        entidad_cuit: CUIT del cliente o proveedor
        tipo: 'cliente' o 'proveedor'
        movimientos: Lista de diccionarios con los movimientos
        fecha_desde: Fecha de inicio del período (opcional)
        fecha_hasta: Fecha de fin del período (opcional)
        saldo_final: Saldo final de la cuenta
    
    Returns:
        Bytes del Excel generado
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    
    # Colores
    HEADER_BG = "2C5282"
    HEADER_FONT = "FFFFFF"
    TOTAL_BG = "E2E8F0"
    ALTERNATE_BG = "F7FAFC"
    BORDER_COLOR = "C0C0C0"
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    # Header - Información de la entidad
    ws.merge_cells('A1:E1')
    ws['A1'] = "AYMARA CONTABLE"
    ws['A1'].font = Font(bold=True, size=16, color=HEADER_FONT)
    ws['A1'].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = "Estado de Cuenta"
    ws['A2'].font = Font(bold=True, size=14, color="2C5282")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Tipo de entidad
    tipo_label = "Cliente" if tipo == 'cliente' else "Proveedor"
    ws['A4'] = f"{tipo_label}:"
    ws['A4'].font = Font(bold=True)
    ws['B4'] = entidad_nombre
    ws['B4'].font = Font(italic=True)
    
    ws['A5'] = "CUIT:"
    ws['A5'].font = Font(bold=True)
    ws['B5'] = entidad_cuit or "N/A"
    
    # Período
    ws['A6'] = "Período:"
    ws['A6'].font = Font(bold=True)
    if fecha_desde and fecha_hasta:
        ws['B6'] = f"{fecha_desde} al {fecha_hasta}"
    else:
        ws['B6'] = "Todo el histórico"

    # Espacio
    ws.row_dimensions[8].height = 15
    
    # Headers de la tabla
    headers = ['Fecha', 'Descripción', 'Debe', 'Haber', 'Saldo']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Movimientos
    row_num = 10
    for idx, mov in enumerate(movimientos):
        # Fecha
        fecha_str = mov['fecha'][:10] if mov['fecha'] else ''
        ws.cell(row=row_num, column=1, value=fecha_str).alignment = Alignment(horizontal='center')
        
        # Descripción
        ws.cell(row=row_num, column=2, value=mov['descripcion'] or '')
        
        # Debe
        debe_val = mov['debe'] if mov['debe'] > 0 else None
        ws.cell(row=row_num, column=3, value=debe_val).number_format = '$#,##0.00'
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right')
        
        # Haber
        haber_val = mov['haber'] if mov['haber'] > 0 else None
        ws.cell(row=row_num, column=4, value=haber_val).number_format = '$#,##0.00'
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
        
        # Saldo
        ws.cell(row=row_num, column=5, value=mov['saldo']).number_format = '$#,##0.00'
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
        
        # Aplicar bordes y formato a toda la fila
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            
            # Filas alternadas
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color=ALTERNATE_BG, end_color=ALTERNATE_BG, fill_type='solid')
        
        row_num += 1
    
    # Fila de totales
    total_debe = sum(m['debe'] for m in movimientos)
    total_haber = sum(m['haber'] for m in movimientos)
    
    ws.cell(row=row_num, column=1, value='')
    ws.cell(row=row_num, column=2, value='TOTALES')
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    ws.cell(row=row_num, column=3, value=total_debe).number_format = '$#,##0.00'
    ws.cell(row=row_num, column=3).font = Font(bold=True)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=4, value=total_haber).number_format = '$#,##0.00'
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=5, value=saldo_final).number_format = '$#,##0.00'
    ws.cell(row=row_num, column=5).font = Font(bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
    
    # Formato especial para fila de totales
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type='solid')
        cell.border = thin_border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    
    # Footer
    row_num += 2
    ws.merge_cells(f'A{row_num}:E{row_num}')
    footer_cell = ws.cell(row=row_num, column=1, value=f"Documento emitido el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    footer_cell.font = Font(italic=True, size=9, color="718096")
    footer_cell.alignment = Alignment(horizontal='center')
    
    row_num += 1
    ws.merge_cells(f'A{row_num}:E{row_num}')
    footer_cell2 = ws.cell(row=row_num, column=1, value="Aymara Contable - Sistema de Gestión Comercial")
    footer_cell2.font = Font(italic=True, size=9, color="718096")
    footer_cell2.alignment = Alignment(horizontal='center')
    
    # Guardar Excel
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes
